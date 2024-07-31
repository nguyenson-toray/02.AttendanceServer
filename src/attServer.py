# -*- coding: utf-8 -*-
import sys
import time
import threading
import os
import schedule
from zk import ZK
import pymongo
from datetime import datetime, timedelta
import pandas as pd
import cv2
from pyzbar.pyzbar import decode
from pdf2image import convert_from_path
from openpyxl import load_workbook
import pyodbc

CWD = os.path.dirname(os.path.realpath(__file__))
ROOT_DIR = os.path.dirname(CWD)
sys.path.append(ROOT_DIR)

client = pymongo.MongoClient("mongodb://localhost:27017/")
db = client["tiqn"]
collection_employee = db["Employee"]
collection_att_log = db["AttLog"]
collection_maternity_tracking = db["MaternityTracking"]
collection_history_get_att_logs = db["HistoryGetAttLogs"]
collection_ot_register = db["OtRegister"]
list_emp = []
finger_emp_id_access_db = {}  #{"emp_id": 0001}
ip_fs = r'\\192.168.1.13'
ip_att_machines = []
path_config = {}


def write_log(log):
    log_str = str(log)
    try:
        file_name = r"..\03.Logs\log_" + datetime.now().strftime("%Y%m%d") + ".txt"
        with open(file_name, 'a', encoding="utf-8") as file:
            file.writelines(log_str)
            file.close()
    except Exception as e:
        print(f'!!!!!!!!! write_log Exception :{e}')


def read_access_db_hr(path: str):
    log = f'{datetime.now()} read_access_db_hr: {path}\n'
    print(log)
    # print('pyodbc.drivers')
    # for x in pyodbc.drivers():
    #     print(x)
    finger_emp_id_access_db.clear()
    conn_str = f"Driver=Microsoft Access Driver (*.mdb, *.accdb);DBQ={path}"
    # print(conn_str)
    conn = pyodbc.connect(conn_str)
    curs = conn.cursor()
    sql = 'SELECT Badgenumber, SSN FROM USERINFO;'
    try:
        curs.execute(sql)
        rows = curs.fetchall()
        curs.close()
        conn.close()
        for row in rows:
            # '1267', 'TIQN-1208'
            row_1 = str(row)[1:-1]
            finger_id = row_1.split(', ')[0][1:-1]  # remove first & last character '
            emp_id = row_1.split(', ')[1][1:-1]  # remove first & last character '
            finger_emp_id_access_db[emp_id] = int(finger_id)
            log += f'    {emp_id}    {finger_id}; '
    except Exception as e:
        log += e
        print(f'    {e}')
    write_log(log+'\n')


def excel_aio_to_db():
    # Read data from Excel using pandas
    log = f'{datetime.now()} excel_aio_to_db\n'
    print(log)
    excel_file = path_config['aio']
    data = pd.read_excel(excel_file, keep_default_na=False, na_values='', na_filter=False)
    data.fillna("", inplace=True)
    # Convert pandas dataframe to dictionary (adjust based on your data structure)
    data_dict = data.to_dict(orient="records")
    # Loop through each data row
    count = 0
    for row in data_dict:
        # Update document in MongoDB collection based on a unique identifier (replace with your logic)
        if row["Emp Code"] == '' or row["Emp Code"] == 0 or row["Fullname"] == '' or row["Fullname"] == 0:
            # print(f'BYPASS ROW - Empty Emp Code or Fullname')
            continue
        if row["Fullname"] == 'Shoji Izumi' or row["Fullname"] == 'Amagata Osamu':
            count += 1
            log += f'   BYPASS Izumi, Amagata\n'
            continue
        count += 1
        empId = 'TIQN-XXXX' if row['Emp Code'] == '' else row['Emp Code']
        filter = {"empId": empId}  # Assuming "_id" is a unique identifier in your data
        fieldCollected = {}
        name = 'No name' if row['Fullname'] == '' else row['Fullname']
        attFingerId = find_finger_id_access_db(empId)
        department = '' if row['Department'] == '' else row['Department']
        section = '' if row['Section'] == '' else row['Section']
        group = '' if row['Group'] == '' else row['Group']
        lineTeam = '' if row['Line/ Team'] == '' else row['Line/ Team']
        gender = '' if row['Gender'] == '' else row['Gender']
        position = '' if row['Position'] == '' else row['Position']
        level = '' if row['Level'] == '' else row['Level']
        directIndirect = '' if row['Direct/ Indirect'] == '' else row['Direct/ Indirect']
        sewingNonSewing = '' if row['Sewing/Non sewing'] == '' else row['Sewing/Non sewing']
        supporting = '' if row['Supporting'] == '' else row['Supporting']
        dob = datetime.fromisoformat('1900-01-01') if str(type(row['DOB'])).find('datetime.datetime') == -1 else row[
            'DOB']
        joiningDate = datetime.fromisoformat('1900-01-01') if str(type(row['Joining date'])).find(
            'datetime.datetime') == -1 else row['Joining date']
        workStatus = 'Resigned' if (row['Working/Resigned'] == 0 or row['Working/Resigned'] == '') else 'Working'
        resignOn = datetime.fromisoformat('2099-01-01')
        fieldCollected.update(
            {'empId': empId, 'name': name, 'attFingerId': attFingerId, 'department': department, 'section': section,
             'group': group, 'lineTeam': lineTeam, 'gender': gender, 'position': position, 'level': level,
             'directIndirect': directIndirect, 'sewingNonSewing': sewingNonSewing, 'supporting': supporting, 'dob': dob,
             'joiningDate': joiningDate, 'workStatus': workStatus, 'resignOn': resignOn})
        update = {"$set": fieldCollected}
        log += f'   update #{count}: {empId}     {name}\n'
        # print(f"collection.update_one: {update}")
        collection_employee.update_one(filter, update, upsert=True)  # Upsert inserts if no document matches the filter
    log += f"    {count} records\n"
    print(f"    {count} records") if enable_print else None
    write_log(log)


def excel_maternity_to_db():
    log = f'{datetime.now()} excel_maternity_to_db\n'
    print(log)
    excel_file = path_config['maternity']
    # Read -thai san
    data = pd.read_excel(excel_file, sheet_name='Thai sản', keep_default_na=False, na_values='',
                         na_filter=False, skiprows=3)
    data.fillna("", inplace=True)
    # Convert pandas dataframe to dictionary (adjust based on your data structure)
    data_dict = data.to_dict(orient="records")
    # Loop through each data row
    empIdMaternity = 'TIQN-XXXX'
    for row in data_dict:
        if row['STT'] != 0 and row['STT'] != '':
            empIdMaternity = row['MSNV']
            query = {'empId': empIdMaternity}
            maternityLeaveBegin = datetime.fromisoformat('2099-01-01') if (str(type(row['NGÀY NGHỈ SINH'])).find(
                'datetime.datetime') == -1 and str(type(row['NGÀY NGHỈ SINH'])).find(
                'timestamps.Timestamp') == -1) else row['NGÀY NGHỈ SINH']
            maternityLeaveEnd = datetime.fromisoformat('2099-01-01') if (str(type(row['NGÀY QUAY LẠI'])).find(
                'datetime.datetime') == -1 and str(type(row['NGÀY QUAY LẠI'])).find(
                'timestamps.Timestamp') == -1) else row['NGÀY QUAY LẠI']
            maternityLeaveEnd = maternityLeaveEnd - timedelta(days=1)
            if (maternityLeaveBegin.year != 2099 and maternityLeaveEnd.year != 2099):
                update = {"$set": {'workStatus': 'Maternity leave', 'maternityLeaveBegin': maternityLeaveBegin,
                                   'maternityLeaveEnd': maternityLeaveEnd}}
            else:
                update = {"$set": {'workStatus': 'Maternity leave'}}
            log += f'   update {empIdMaternity} to Maternity leave\n'
            print(f"    update {empIdMaternity} to Maternity leave") if enable_print else None
            collection_employee.update_one(query, update)

    ##---------mang thai--pregnant-------
    data = pd.read_excel(excel_file, sheet_name='mang thai', keep_default_na=False, na_values='',
                         na_filter=False, skiprows=3)
    data.fillna("", inplace=True)
    # Convert pandas dataframe to dictionary (adjust based on your data structure)
    data_dict = data.to_dict(orient="records")
    # Loop through each data row
    for row in data_dict:
        if row['STT'] != 0 and row['STT'] != '':
            empIdMaternity = row['MSNV']
            maternityBegin = row['NGÀY NHẬN THÔNG TIN']
            maternityEnd = row['NGÀY DỰ SINH']
            query = {'empId': empIdMaternity}
            update = {"$set": {'workStatus': 'Working pregnant', 'maternityBegin': maternityBegin,
                               'maternityEnd': maternityEnd}}
            collection_employee.update_one(query, update)
            log += f"   update {empIdMaternity} to Working pregnant\n"
            print(f"    update {empIdMaternity} to Working pregnant") if enable_print else None
    ##---------Con nhỏ dưới 12 tháng---------
    data = pd.read_excel(excel_file, sheet_name='Con nhỏ dưới 12 tháng', keep_default_na=False, na_values='',
                         na_filter=False, skiprows=3)
    data.fillna("", inplace=True)
    # Convert pandas dataframe to dictionary (adjust based on your data structure)
    data_dict = data.to_dict(orient="records")
    # Loop through each data row
    for row in data_dict:
        if row['STT'] != 0 and row['STT'] != '':
            empIdMaternity = row['MSNV']
            maternityBegin = row['NGÀY QUAY LẠI']
            maternityEnd = row['NGÀY CUỐI CÙNG THỜI GIAN NUÔI CON NHỎ']
            query = {'empId': empIdMaternity}
            update = {"$set": {'workStatus': 'Working young child', 'maternityBegin': maternityBegin,
                               'maternityEnd': maternityEnd}}
            log += f"   update {empIdMaternity} to Working young child\n"
            print(f"   update {empIdMaternity} to Working young child") if enable_print else None
            collection_employee.update_one(query, update)
    write_log(log)


def excel_resign_to_db():
    log = f'{datetime.now()} excel_resign_to_db\n'
    print(log)
    excel_file = path_config['resign']
    data = pd.read_excel(excel_file, sheet_name='QD', keep_default_na=False, na_values='',
                         na_filter=False)
    data.fillna("", inplace=True)
    # Convert pandas dataframe to dictionary (adjust based on your data structure)
    data_dict = data.to_dict(orient="records")
    # Loop through each data row
    for row in data_dict:
        if row['Số QĐ'] != 0 and row['Số QĐ'] != '':
            empIdResigned = row['MSNV']
            resignDate = datetime.fromisoformat('2099-01-01') if (str(type(row['Ngày nghỉ việc'])).find(
                'datetime.datetime') == -1 and str(type(row['Ngày nghỉ việc'])).find(
                'timestamps.Timestamp') == -1) else row['Ngày nghỉ việc']

            if resignDate.year > 1900 and resignDate < datetime.now():
                query = {'empId': empIdResigned}
                update = {"$set": {'workStatus': 'Resigned', 'resignOn': resignDate}}
                log += f"    update {empIdResigned} to resigned on {resignDate}\n"
                print(f"    update {empIdResigned} to resigned on {resignDate}") if enable_print else None
                collection_employee.update_one(query, update)
    write_log(log)


def get_att_log_one_time(machine: ZK, machineNo: int) -> int:
    conn = None
    count = 0
    log = f'{datetime.now()} get_att_log_one_time: machine {machineNo}\n'
    print(f'{datetime.now()} get_att_log_one_time: machine {machineNo}')
    time_begin_get_logs = datetime.now()
    try:

        history = {}
        for his in collection_history_get_att_logs.find():
            if int(his['machine']) == machineNo:
                history = his
        last_time_db = history['lastTimeGetAttLogs']
        last_time_machine = last_time_db
        conn = machine.connect()
        print(f"Connecting machine : {machineNo} ...")
        # disable device, this method ensures no activity on the device while the process is run
        conn.disable_device()
        # another commands will be here!
        # Get attendances (will return list of Attendance object)
        attendances = conn.get_attendance()
        for attendance in attendances:
            if attendance.timestamp > last_time_db:
                mydict = {"machineNo": machineNo, "uid": attendance.uid, "attFingerId": int(attendance.user_id),
                          "empId": 'No Emp Id', "name": 'No name',
                          "timestamp": attendance.timestamp}
                emp_id = find_emp_id_by_finger_id(int(attendance.user_id))
                emp_name = find_name_by_finger_id(int(attendance.user_id))
                if emp_id != 'Not found':
                    mydict['empId'] = emp_id
                if emp_name != 'Not found':
                    mydict['name'] = emp_name
                log1 = f'    Machine: {mydict['machineNo']}       {mydict['timestamp'].strftime("%d-%m-%Y %H:%M:%S")}      {mydict['attFingerId']}     {mydict['empId']}    {mydict['name']}\n'
                collection_att_log.insert_one(mydict)
                print(log1)
                log += log1
                count += 1
                last_time_machine = attendance.timestamp
        my_query = {"machine": machineNo}
        new_value = {"$set": {"lastTimeGetAttLogs": last_time_machine, "lastCount": count}}
        collection_history_get_att_logs.update_one(my_query, new_value)
        conn.enable_device()
    except Exception as e:
        print(f'     get_att_log_one_time exception: {e}')
        write_log(f'     get_att_log_one_time exception: {e}')
    finally:
        if conn:
            conn.disconnect()
        log += f'    Machine {machineNo} => {count} records. Total time: {datetime.now() - time_begin_get_logs}\n'
        print(
            f"    Machine {machineNo} => {count} records. Total time: {datetime.now() - time_begin_get_logs}")
    write_log(log)


def live_capture_attendance(machine: ZK, machineNo) -> None:
    conn = None
    log = f'{datetime.now()} live_capture_attendance: machine {machineNo}\n'
    write_log(log)
    print(f'{datetime.now()} live_capture_attendance: machine {machineNo}')
    try:
        conn = machine.connect()
        for attendance in conn.live_capture():
            if attendance is None:
                pass
            else:
                mydict = {"machineNo": machineNo, "uid": attendance.uid, "attFingerId": int(attendance.user_id),
                          "empId": 'No Emp Id', "name": 'No name',
                          "timestamp": attendance.timestamp}

                emp_id = find_emp_id_by_finger_id(int(attendance.user_id))
                emp_name = find_name_by_finger_id(int(attendance.user_id))
                if emp_id != 'Not found':
                    mydict['empId'] = emp_id
                if emp_name != 'Not found':
                    mydict['name'] = emp_name
                log = f'    Live capture - Machine: {mydict['machineNo']}      {mydict['timestamp'].strftime("%d-%m-%Y %H:%M:%S")}     {mydict['attFingerId']}    {mydict['empId']}   {mydict['name']}\n'
                collection_att_log.insert_one(mydict)
                print(log)
                write_log(log)
                myquery = {"machine": machineNo}
                new_value = {"$set": {"lastTimeGetAttLogs": attendance.timestamp, "lastCount": 1}}
                collection_history_get_att_logs.update_one(myquery, new_value)

    except Exception as e:
        print("     Exception: Process terminate : {}".format(e))
        write_log(f'    live_capture Exception: {e}')
    finally:
        if conn:
            conn.disconnect()


def update_excel_to_mongoDb():
    excel_aio_to_db()
    excel_maternity_to_db()
    excel_resign_to_db()
    get_list_emp()


def get_list_emp():
    list_emp.clear()
    count = 0
    for emp in collection_employee.find():
        count += 1
        list_emp.append({'attFingerId': emp['attFingerId'], 'empId': emp['empId'], 'name': emp['name']})
    print(f'{datetime.now()} get_list_emp: Total:{count}\n') if enable_print else None
    log = f'{datetime.now()} get_list_emp:\n{list_emp} \nTotal:{count}\n'

    excel_file = r"..\02.Config\config.xlsx"
    print(f'{datetime.now()} read_config : {excel_file}') if enable_print else None
    data_phu_quy = pd.read_excel(excel_file, sheet_name='phú quý')
    data_dict_phu_quy = data_phu_quy.to_dict(orient="records")
    for row in data_dict_phu_quy:
        list_emp.append({'attFingerId': row['FingerID'], 'empId': row['Code'], 'name': row['Name']})
    write_log(log)


def ot_register_detect_qr_and_save():
    log = ''
    ot_folder_path = path_config['ot_folder']
    ot_folder_pdf_imported_path = ot_folder_path + r"\01.Imported"
    ot_folder_pdf_path = ot_folder_path + r"\02.Pdf"
    ot_excel_file_name = "01.OT summary.xlsx"
    my_poppler_path = r'C:\Program Files\poppler-24.02.0\Library\bin'
    ot_request_no_on_db = []
    for doc in collection_ot_register.find():
        ot_request_no_on_db.append(str(doc['requestNo']))

    ot_last_request_id_db = 0
    ot_last_ot_register_document = collection_ot_register.find().sort({'_id': -1}).limit(1)
    for doc in ot_last_ot_register_document:
        ot_last_request_id_db = doc['_id']
    pdf_files = []
    try:
        # Use glob with wildcard for efficient PDF search
        for filename in os.listdir(ot_folder_pdf_path):
            if filename.endswith('.pdf'):  # and filename.startswith("202"):
                pdf_files.append(filename)
    except Exception as e:
        log += f'   Exception : Error processing folder {ot_folder_pdf_path}: {e}\n'
        print(f"    Exception : Error processing folder {ot_folder_pdf_path}: {e}")
    for file_pdf in pdf_files:
        file_path = os.path.join(ot_folder_pdf_path, file_pdf)
        log += f'    Read QR code in file: {file_path}\n'
        print(f'    Read QR code in file: {file_path}') if enable_print else None
        try:
            temp_pages = convert_from_path(file_path, dpi=300, output_file='qr.png', paths_only=True,
                                           output_folder=ot_folder_pdf_path, poppler_path=my_poppler_path)
            for temp_page in temp_pages:
                try:
                    page = cv2.imread(temp_page)
                    gray = cv2.cvtColor(page, cv2.COLOR_BGR2GRAY)  # Convert to grayscale for better detection
                    # Detect QR codes
                    qr_codes = decode(gray)
                    # Extract data from QR codes
                    for qr_code in qr_codes:
                        qr = str(qr_code.data.decode('utf-8'))
                        print(f'File : {file_pdf} :    Detected QR code: {qr}\n') if enable_print else None
                        log += f'File : {file_pdf} :    Detected QR code: {qr}\n'
                        ot_request_no = qr.split(';')[0].strip()
                        date_request = qr.split(';')[1].strip()
                        list_date_time = qr.split(';')[2].strip().split(', ')
                        list_emp_id = qr.split(';')[3].strip().split(' ')
                        if (not ot_request_no_on_db.__contains__(ot_request_no)):
                            ot_request_no_on_db.append(ot_request_no)
                            ot_last_request_id_db = qr_code_ot_register_to_db(ot_last_request_id_db, ot_request_no,
                                                                              date_request,
                                                                              list_date_time, list_emp_id)
                            print(f'    Add QR to DB: {qr}') if enable_print else None
                            log += f'    Add QR to DB: {qr}\n'
                        else:
                            print(f'    Duplicate QR on DB: {qr}') if enable_print else None
                            log += f'    Duplicate QR on DB:: {qr}\n'
                except Exception as e:
                    log += f'   detect_qr : Error 1 :{e}\n'
                    print(f"    detect_qr : Error 1 :{e}")

        except Exception as e:
            log += f'   detect_qr : Error 2 :{e}\n'
            print(f"    detect_qr : Error 2 :{e}")
        src_path = os.path.join(ot_folder_pdf_path, file_pdf)
        dst_path = os.path.join(ot_folder_pdf_imported_path, file_pdf)
        os.rename(src_path, dst_path)  # move file pdf scaned to "01.Imported"
        log += f'   Scanned & moved to folder "01.Imported" : {file_pdf}'
        print(f'    Scanned & moved to folder "01.Imported" : {file_pdf}') if enable_print else None
        for filename in os.listdir(ot_folder_pdf_path):
            file_path = os.path.join(ot_folder_pdf_path, filename)
            if filename.endswith('.ppm'):
                os.remove(file_path)
                print(f'    Delete .ppm : {filename}') if enable_print else None

    try:
        ot_register_append_excel(ot_folder_path, ot_excel_file_name)
    except Exception as e:
        log += f'   detect_qr Exception: Error 3 :{e}\n'
        print(f"    detect_qr Exception: Error 3 :{e}")
    write_log(log)


def ot_register_append_excel(folder_path: str, file_name: str):
    filename = os.path.join(folder_path, file_name)
    ot_last_request_id_db = 0
    ot_last_ot_register_document = collection_ot_register.find().sort({'_id': -1}).limit(1)
    for doc in ot_last_ot_register_document:
        ot_last_request_id_db = doc['_id']
    if ot_last_request_id_db == 0:
        return
    # Open the existing workbook
    wb = load_workbook(filename=filename)
    sheet_name = "data"  # Replace with your sheet name
    # Select the sheet you want to append data to (adjust sheet name)
    ws = wb[sheet_name]

    if ws.max_row == 1:
        ot_last_request_id_excel = 0
        last_row = 1
    else:
        last_row = ws.max_row
        ot_last_request_id_excel = int(ws.cell(last_row, 1).value)
    if ot_last_request_id_db > ot_last_request_id_excel:
        log = f'{datetime.now()} ot_register_append_excel\n'
        print(F'{datetime.now()} Append OT data from db to excel') if enable_print else None
        query = {"_id": {"$gt": ot_last_request_id_excel}}
        docs = collection_ot_register.find(query)
        for doc in docs:
            log += f"   Append ID : {doc['_id']}\n"
            print(f"   Append ID : {doc['_id']}\n") if enable_print else None
            last_row += 1
            ws.cell(row=last_row, column=1).value = doc['_id']
            ws.cell(row=last_row, column=2).value = doc['requestNo']
            ws.cell(row=last_row, column=3).value = doc['requestDate']
            ws.cell(row=last_row, column=4).value = doc['otDate']
            ws.cell(row=last_row, column=5).value = doc['otTimeBegin']
            ws.cell(row=last_row, column=6).value = doc['otTimeEnd']
            ws.cell(row=last_row, column=7).value = doc['empId']
            ws.cell(row=last_row, column=8).value = doc['name']
        # Apply date formatting to the entire column
        date_format = "DD-MM-YYYY"
        for row in ws.iter_rows(min_row=2):  # Skip the first row (headers)
            cell = row[2]  # Access cell by column index (0-based)
            if isinstance(cell.value, datetime):  # Check if value is a date object
                cell.number_format = date_format
            else:
                pass
            cell = row[3]  # Access cell by column index (0-based)
            if isinstance(cell.value, datetime):  # Check if value is a date object
                cell.number_format = date_format
            else:
                pass
        # Save the modified workbook
        wb.save(filename=filename)
        wb.close()
        print(f"  => ot_register_append_excel : Data appended to: {filename}") if enable_print else None
        write_log(log)


def qr_code_ot_register_to_db(ot_last_request_id_db: int, ot_request_no: str, date_request: str, list_date_time: list,
                              list_emp_id: list) -> int:
    log = f'{datetime.now()} qr_code_ot_register_to_db: {ot_request_no}     List Date & Time: {list_date_time} List Emp ID: {list_emp_id}\n'
    print(
        f'{datetime.now()} qr_code_ot_register_to_db: {ot_request_no}     List Date & Time: {list_date_time} List Emp ID: {list_emp_id}\n') if enable_print else None
    for date_and_time in list_date_time:
        if not len(date_and_time) == 20 or date_and_time.find('19000100') >= 0:
            print(f'{date_and_time}: is wrong => pass') if enable_print else None
            continue
        date_ot_str = date_and_time.strip().split(' ')[0]
        ot_time_begin_str = date_and_time.strip().split(' ')[1]
        ot_time_end_str = date_and_time.strip().split(' ')[2]
        ot_date = datetime.strptime(date_ot_str[0:8], '%Y%m%d')
        request_date = datetime.strptime(date_request, '%Y%m%d')
        for id_no in list_emp_id:
            emp_id = 'TIQN-' + str(id_no)
            emp_name = find_name_by_emp_id(emp_id)
            ot_last_request_id_db += 1
            mydict = {"_id": ot_last_request_id_db, "requestNo": ot_request_no, "requestDate": request_date,
                      "otDate": ot_date, 'otTimeBegin': ot_time_begin_str, 'otTimeEnd': ot_time_end_str,
                      "empId": emp_id, "name": emp_name}
            log += f'   Add record OT register: No: {ot_request_no}      Request date: {mydict['requestDate']}      Date: {mydict['otDate']}       From: {mydict['otTimeBegin']}       To: {mydict['otTimeEnd']}       Emp ID: {mydict['empId']}       Name: {mydict['name']}\n'
            collection_ot_register.insert_one(mydict)
    write_log(log)
    return ot_last_request_id_db


def find_name_by_emp_id(emp_id: str) -> str:
    name = 'Not found'
    for emp in list_emp:
        if emp['empId'] == emp_id:
            name = emp['name']
    return name


def find_name_by_finger_id(finger_id: int) -> str:
    name = 'Not found'
    for emp in list_emp:
        if emp['attFingerId'] == finger_id:
            name = emp['name']
            break
    return name


def find_finger_id_access_db(emp_id: str) -> int:
    finger_id = 0
    try:
        finger_id = finger_emp_id_access_db[emp_id]
    except:
        # print(f'find_finger_id_access_db : Not found emp_id : {emp_id}')
        pass
    return finger_id


def find_emp_id_by_finger_id(finger_id: int) -> str:
    emp_id = 'Not found'
    for emp in list_emp:
        if emp['attFingerId'] == finger_id:
            emp_id = emp['empId']
            break
    return emp_id


def get_att_log(machine: ZK, machineNo, real_time: bool):
    get_att_log_one_time(machine, machineNo)
    if real_time:
        live_capture_attendance(machine, machineNo)


def sync_time_devices():
    log = f'{datetime.now()} sync_time_devices\n'
    try:
        for ip in ip_att_machines:
            machine = ZK(ip, port=4370, timeout=5, password=0, force_udp=False, ommit_ping=False)
            conn = machine.connect()
            time = datetime.now()
            print(f"    Machine : {ip}\n    Current time : {conn.get_time()}    Sync time to : {time}")
            conn.set_time(time)
    except Exception as e:
        log += f'   Sync time ERROR: {e}\n'
        print(f"     Sync time ERROR : {e}")
    write_log(log)


def read_config():
    ip_att_machines.clear()
    log = f'{datetime.now()} read_config:\n'
    excel_file = r"..\02.Config\config.xlsx"
    print(f'{datetime.now()} read_config : {excel_file}')
    data_ip = pd.read_excel(excel_file, sheet_name='ip_machines')
    data_dict_ip = data_ip.to_dict(orient="records")
    for row in data_dict_ip:
        ip_att_machines.append(row['IP'])
    log += f'     ip_att_machines: {ip_att_machines}\n'
    print(f'    ip_att_machines: {ip_att_machines}') if enable_print else None
    data_path = pd.read_excel(excel_file, sheet_name='path')
    data_dict_path = data_path.to_dict(orient="records")
    path_config.clear()
    for row in data_dict_path:
        if row['Name'] == 'aio':
            path_config['aio'] = row['Path']
        if row['Name'] == 'resign':
            path_config['resign'] = row['Path']
        if row['Name'] == 'maternity':
            path_config['maternity'] = row['Path']
        if row['Name'] == 'ot_folder':
            path_config['ot_folder'] = row['Path']
        if row['Name'] == 'maternity_pregnant':
            path_config['maternity_pregnant'] = row['Path']
        if row['Name'] == 'maternity_young_child':
            path_config['maternity_young_child'] = row['Path']
        if row['Name'] == 'maternity_leave':
            path_config['maternity_leave'] = row['Path']
        if row['Name'] == 'access_db':
            path_config['access_db'] = row['Path']
    print(f'    path_config: {path_config}') if enable_print else None
    log += f'      path_config: {str(path_config)}\n'
    write_log(log)


if __name__ == "__main__":
    enable_print = False
    main_log = f'---------------------------*** ATTENDANCE SERVER ***-----------------------------------------\n'
    main_log += f'START AT : {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}\n'
    main_log += '''
    schedule.every().sunday.at("05:00").do(sync_time_devices)
    schedule.every().day.at("09:00").do(update_excel_to_mongoDb)
    schedule.every().day.at("11:55").do(update_excel_to_mongoDb)
    schedule.every().day.at("15:00").do(update_excel_to_mongoDb)
    schedule.every().day.at("18:00").do(update_excel_to_mongoDb)
    schedule.every(10).minutes.do(ot_register_detect_qr_and_save)'''
    main_log += f'\n-------------------------------------------------------------------------------------------\n'
    print(main_log)
    write_log(main_log)
    read_config()
    read_access_db_hr(path_config['access_db'])
    update_excel_to_mongoDb()
    get_list_emp()
    machine_no = 1
    for ip in ip_att_machines:
        machine = ZK(ip, port=4370, timeout=5, password=0, force_udp=False, ommit_ping=False)
        threading.Thread(target=get_att_log, args=(
            machine, machine_no, True), name=ip).start()
        machine_no += 1
    ot_register_detect_qr_and_save()
    schedule.every().sunday.at("05:00").do(sync_time_devices)
    schedule.every().day.at("09:00").do(update_excel_to_mongoDb)
    schedule.every().day.at("11:50").do(update_excel_to_mongoDb)
    schedule.every().day.at("15:00").do(update_excel_to_mongoDb)
    schedule.every().day.at("18:00").do(update_excel_to_mongoDb)
    schedule.every(10).minutes.do(ot_register_detect_qr_and_save)
    while True:
        schedule.run_pending()
        time.sleep(1)
