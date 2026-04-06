# -*- coding: utf-8 -*-
"""
functions.py — All business logic for Attendance Server V2
Imported by attServerV2.py (CLI) and attServerV2UI.py (GUI)
"""
import sys
import time as time_module
import threading
import os
import json
import schedule
import numpy as np
from zk import ZK
import pymongo
from pymongo import UpdateOne
from datetime import datetime, timedelta
import pandas as pd
import cv2
from pyzbar.pyzbar import decode as pyzbar_decode
from pdf2image import convert_from_path
from openpyxl import load_workbook

# optional: zxingcpp — pip install zxingcpp
try:
    import zxingcpp
    _HAS_ZXINGCPP = True
except ImportError:
    _HAS_ZXINGCPP = False

# ─── Constants ────────────────────────────────────────────────────────────────
ATT_LOG_INTERVAL_MINUTES = 6       # default fallback if not in config
REAL_TIME = False

LIVE_CAPTURE_TIMEOUT = 10
ZOMBIE_THRESHOLD = 6
RECONNECT_DELAY_MIN = 10
RECONNECT_DELAY_MAX = 300
RECONNECT_DELAY_MULTIPLIER = 2

QR_PDF_DPI = 400   # higher than 300 → better quality for blurry scans

# ─── Paths ────────────────────────────────────────────────────────────────────
CWD = os.path.dirname(os.path.realpath(__file__))
ROOT_DIR = os.path.dirname(CWD)
sys.path.append(ROOT_DIR)

CONFIG_PATH = os.path.join(ROOT_DIR, '02.Config', 'config.json')
LOG_DIR     = os.path.join(ROOT_DIR, '03.Logs')

# ─── Logging ──────────────────────────────────────────────────────────────────
enable_print = False

os.makedirs(LOG_DIR, exist_ok=True)
_log_lock = threading.Lock()


def _write_log(level: str, message: str) -> None:
    log_str   = f'{datetime.now():%d-%m-%Y %H:%M:%S} {message}\n'
    file_path = os.path.join(LOG_DIR, f'{level}_{datetime.now():%Y%m%d}.txt')
    try:
        with _log_lock:
            with open(file_path, 'a', encoding='utf-8') as f:
                f.write(log_str)
    except Exception as e:
        print(f'!!! _write_log Exception: {e}')


def write_log_info(message: str, module_name: str = 'GENERAL') -> None:
    msg = f'[{module_name}] INFO: {message}'
    _write_log('info', msg)
    if enable_print:
        print(msg)


def write_log_error(error_message: str, module_name: str = 'GENERAL', exception=None) -> None:
    msg = f'[{module_name}] ERROR: {error_message}'
    if exception:
        msg += f' | Exception: {exception}'
    _write_log('error', msg)
    if enable_print:
        print(msg)


def write_log_debug(message: str, module_name: str = 'GENERAL') -> None:
    _write_log('debug', f'[{module_name}] DEBUG: {message}')


# ─── MongoDB ──────────────────────────────────────────────────────────────────
client = pymongo.MongoClient("mongodb://localhost:27017/")
db = client["tiqn"]
collection_employee              = db["Employee"]
collection_att_log               = db["AttLog"]
collection_maternity_tracking    = db["MaternityTracking"]
collection_history_get_att_logs  = db["HistoryGetAttLogs"]
collection_ot_register           = db["OtRegister"]

# ─── Global state ─────────────────────────────────────────────────────────────
ip_att_machines: list = []
path_config:     dict = {}
sheet_config:    dict = {}
schedule_config: dict = {}
bypass_names:    list = []
poppler_path:    str  = ''

emp_by_finger_id: dict = {}   # {attFingerId: {'empId': str, 'name': str}}
emp_by_emp_id:    dict = {}   # {empId: {'attFingerId': int, 'name': str}}

# update window — only get_att_log & update_excel run inside [from, to]
update_time_from: str = '17:00'
update_time_to:   str = '22:00'


# ─── Config ───────────────────────────────────────────────────────────────────
def is_in_update_window() -> bool:
    """Return True if current time is within [update_time_from, update_time_to]."""
    now   = datetime.now().strftime('%H:%M')
    return update_time_from <= now <= update_time_to


def read_config() -> None:
    global poppler_path, update_time_from, update_time_to
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        cfg = json.load(f)
    ip_att_machines.clear()
    ip_att_machines.extend(cfg['att_machines'])
    path_config.clear()
    path_config.update(cfg['paths'])
    sheet_config.clear()
    sheet_config.update(cfg['sheets'])
    schedule_config.clear()
    schedule_config.update(cfg.get('schedule', {
        'sync_time_day':            'sunday',
        'sync_time_at':             '06:00',
        'excel_sync_times':         ['07:00', '09:00', '11:50', '15:00', '18:00', '21:00'],
        'ot_scan_interval_minutes':  10,
        'att_log_interval_minutes':  6,
    }))
    bypass_names.clear()
    bypass_names.extend(cfg.get('bypass_names', []))
    poppler_path      = cfg.get('poppler_path', '')
    update_time_from  = schedule_config.get('update_time_from', '05:00')
    update_time_to    = schedule_config.get('update_time_to',   '22:00')
    write_log_info(f'att_machines: {ip_att_machines}', 'CONFIG')
    write_log_info(f'paths: {path_config}', 'CONFIG')
    write_log_info(f'sheets: {sheet_config}', 'CONFIG')
    write_log_info(f'schedule: {schedule_config}', 'CONFIG')
    write_log_info(f'update_window: {update_time_from} → {update_time_to}', 'CONFIG')


# ─── Employee lookup ──────────────────────────────────────────────────────────
def get_list_emp() -> None:
    emp_by_finger_id.clear()
    emp_by_emp_id.clear()
    count = 0
    for emp in collection_employee.find():
        fid = emp['attFingerId']
        emp_by_finger_id[fid] = {'empId': emp['empId'], 'name': emp['name']}
        emp_by_emp_id[emp['empId']] = {'attFingerId': fid, 'name': emp['name']}
        count += 1
    write_log_info(f'get_list_emp: Total {count}', 'EMP')


def find_name_by_finger_id(finger_id: int) -> str:
    return emp_by_finger_id.get(finger_id, {}).get('name', 'Not found')

def find_emp_id_by_finger_id(finger_id: int) -> str:
    return emp_by_finger_id.get(finger_id, {}).get('empId', 'Not found')

def find_name_by_emp_id(emp_id: str) -> str:
    return emp_by_emp_id.get(emp_id, {}).get('name', 'Not found')


# ─── Excel → MongoDB sync ─────────────────────────────────────────────────────
def excel_aio_to_db() -> None:
    write_log_info('excel_aio_to_db start', 'AIO')
    data = pd.read_excel(path_config['aio'], sheet_name=sheet_config['aio'],
                         keep_default_na=False, na_values='', na_filter=False)
    data.fillna("", inplace=True)
    ops = []
    count = 0
    for row in data.to_dict(orient="records"):
        if (row["Emp Code"] in ('', 0) or row["Fullname"] in ('', 0)
                or row['_id'] in (0, '')):
            write_log_info(f'BYPASS ROW: {row}', 'AIO')
            continue
        if row["Fullname"] in bypass_names:
            count += 1
            continue
        count += 1
        dob          = row['DOB']          if isinstance(row['DOB'],          datetime) else datetime.fromisoformat('1900-01-01')
        joining_date = row['Joining date'] if isinstance(row['Joining date'], datetime) else datetime.fromisoformat('1900-01-01')
        group = row['Group'] if (row.get('Group') and row['Group'] != 0) else ''
        field = {
            'empId':           row['Emp Code'] or 'TIQN-XXXX',
            'name':            row['Fullname'] or 'No name',
            'attFingerId':     0 if row['Finger Id'] == '' else row['Finger Id'],
            'department':      row.get('Department', ''),
            'section':         row.get('Section', ''),
            'group':           group,
            'lineTeam':        group,
            'gender':          row.get('Gender', ''),
            'position':        row.get('Position', ''),
            'level':           row.get('Level', ''),
            'directIndirect':  row.get('Direct/ Indirect', ''),
            'sewingNonSewing': '',
            'supporting':      '',
            'dob':             dob,
            'joiningDate':     joining_date,
            'workStatus':      'Resigned' if row['Working/Resigned'] in (0, '') else 'Working',
            'resignOn':        datetime.fromisoformat('2099-01-01'),
        }
        ops.append(UpdateOne({"_id": row["_id"]}, {"$set": field}, upsert=True))
    if ops:
        collection_employee.bulk_write(ops)
    write_log_info(f'excel_aio_to_db: updated {count} records', 'AIO')


def excel_maternity_to_db() -> None:
    write_log_info('excel_maternity_to_db start', 'MAT')
    excel_file = path_config['maternity']

    # Thai sản
    data = pd.read_excel(excel_file, sheet_name=sheet_config['maternity_leave'],
                         keep_default_na=False, na_values='', na_filter=False, skiprows=3)
    data.fillna("", inplace=True)
    ops, count = [], 0
    for row in data.to_dict(orient="records"):
        if row['STT'] in (0, ''):
            continue
        begin = row['NGÀY NGHỈ SINH'] if isinstance(row['NGÀY NGHỈ SINH'], datetime) else datetime.fromisoformat('2099-01-01')
        end   = row['NGÀY QUAY LẠI']  if isinstance(row['NGÀY QUAY LẠI'],  datetime) else datetime.fromisoformat('2099-01-01')
        end   = end - timedelta(days=1)
        upd   = {'workStatus': 'Maternity leave'}
        if begin.year != 2099 and end.year != 2099:
            upd.update({'maternityLeaveBegin': begin, 'maternityLeaveEnd': end})
        ops.append(UpdateOne({'empId': row['MSNV']}, {"$set": upd}))
        count += 1
    if ops:
        collection_employee.bulk_write(ops)
    write_log_info(f'excel_maternity_to_db: {count} records → Maternity leave', 'MAT')

    # Mang thai
    data = pd.read_excel(excel_file, sheet_name=sheet_config['maternity_pregnant'],
                         keep_default_na=False, na_values='', na_filter=False, skiprows=3)
    data.fillna("", inplace=True)
    ops, count = [], 0
    for row in data.to_dict(orient="records"):
        if (row['STT'] not in (0, '')
                and isinstance(row['NGÀY VỀ CHẾ ĐỘ'], datetime)
                and isinstance(row['NGÀY DỰ SINH'],   datetime)):
            ops.append(UpdateOne(
                {'empId': row['MSNV']},
                {"$set": {'workStatus': 'Working pregnant',
                          'maternityBegin': row['NGÀY VỀ CHẾ ĐỘ'],
                          'maternityEnd':   row['NGÀY DỰ SINH']}}
            ))
            count += 1
        else:
            write_log_error(f'Wrong format at row: {row}', 'MAT')
    if ops:
        collection_employee.bulk_write(ops)
    write_log_info(f'excel_maternity_to_db: {count} records → Working pregnant', 'MAT')

    # Con nhỏ dưới 12 tháng
    data = pd.read_excel(excel_file, sheet_name=sheet_config['maternity_child'],
                         keep_default_na=False, na_values='', na_filter=False, skiprows=3)
    data.fillna("", inplace=True)
    ops, count = [], 0
    for row in data.to_dict(orient="records"):
        if (row['STT'] not in (0, '')
                and isinstance(row['NGÀY QUAY LẠI'], datetime)
                and isinstance(row['NGÀY CUỐI CÙNG THỜI GIAN NUÔI CON NHỎ'], datetime)):
            ops.append(UpdateOne(
                {'empId': row['MSNV']},
                {"$set": {'workStatus': 'Working young child',
                          'maternityBegin': row['NGÀY QUAY LẠI'],
                          'maternityEnd':   row['NGÀY CUỐI CÙNG THỜI GIAN NUÔI CON NHỎ']}}
            ))
            count += 1
        else:
            write_log_error(f'Wrong format at row: {row}', 'MAT')
    if ops:
        collection_employee.bulk_write(ops)
    write_log_info(f'excel_maternity_to_db: {count} records → Working young child', 'MAT')


def excel_resign_to_db() -> None:
    write_log_info('excel_resign_to_db start', 'RESIGN')
    data = pd.read_excel(path_config['resign'], sheet_name=sheet_config['resign'],
                         keep_default_na=False, na_values='', na_filter=False, skiprows=1)
    data.fillna("", inplace=True)
    ops, count = [], 0
    for row in data.to_dict(orient="records"):
        if row['Code'] in (0, ''):
            continue
        resign_on = row['Resign on'] if isinstance(row['Resign on'], datetime) else datetime.fromisoformat('2099-01-01')
        if resign_on.year > 1900 and resign_on <= datetime.now():
            ops.append(UpdateOne(
                {'empId': row['Code']},
                {"$set": {'workStatus': 'Resigned', 'resignOn': resign_on}}
            ))
            count += 1
    if ops:
        collection_employee.bulk_write(ops)
    write_log_info(f'excel_resign_to_db: {count} records → Resigned', 'RESIGN')


def update_excel_to_mongoDb() -> None:
    if not is_in_update_window():
        write_log_info(
            f'update_excel_to_mongoDb — outside update window '
            f'({update_time_from}–{update_time_to}), skip', 'SYNC'
        )
        return
    try:
        excel_aio_to_db()
        excel_maternity_to_db()
        excel_resign_to_db()
        get_list_emp()
    except Exception as e:
        write_log_error(f'update_excel_to_mongoDb: {e}', 'SYNC', e)


# ─── Enhanced QR reader ───────────────────────────────────────────────────────
# Pipelines applied in order. Stop as soon as any pipeline yields a result.
# Each pipeline produces a different preprocessed image to improve decode
# success on blurry / low-contrast / unevenly-lit scans.

_QR_PIPELINES = ('gray', 'otsu', 'adaptive', 'sharpen', 'upscale2x')


def _preprocess(gray: np.ndarray, method: str) -> np.ndarray:
    if method == 'gray':
        return gray
    if method == 'otsu':
        _, out = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        return out
    if method == 'adaptive':
        return cv2.adaptiveThreshold(
            gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
    if method == 'sharpen':
        kernel  = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]], dtype=np.float32)
        sharp   = cv2.filter2D(gray, -1, kernel)
        _, out  = cv2.threshold(sharp, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        return out
    if method == 'upscale2x':
        up     = cv2.resize(gray, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        _, out = cv2.threshold(up, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        return out
    return gray


def _decode_image(img: np.ndarray) -> list:
    """Try pyzbar then cv2.QRCodeDetector on a single image. Return list of data strings."""
    results = {}

    # pyzbar — handles QR + barcodes
    try:
        for code in pyzbar_decode(img):
            data = code.data.decode('utf-8')
            if data:
                results[data] = True
    except Exception:
        pass

    # OpenCV QRCodeDetector — different algorithm, good fallback
    try:
        detector = cv2.QRCodeDetector()
        # detectAndDecodeMulti available in OpenCV 4.5+
        if hasattr(detector, 'detectAndDecodeMulti'):
            ok, data_list, _, _ = detector.detectAndDecodeMulti(img)
            if ok:
                for d in data_list:
                    if d:
                        results[d] = True
        else:
            data, _, _ = detector.detectAndDecode(img)
            if data:
                results[data] = True
    except Exception:
        pass

    return list(results.keys())


def decode_qr_from_image(img_path: str) -> list:
    """
    Robust QR decode: try 5 preprocessing pipelines × 2 decoders.
    Falls back to zxingcpp (if installed) on total failure.

    Returns list of decoded QR data strings (deduplicated).
    Pipeline order (stops as soon as result found):
      1. gray        — baseline, fast
      2. otsu        — high contrast / noisy
      3. adaptive    — uneven lighting
      4. sharpen     — blurry / soft focus
      5. upscale2x   — small / low-res QR code
    """
    page = cv2.imread(img_path)
    if page is None:
        write_log_error(f'decode_qr_from_image: cannot read image {img_path}', 'QR')
        return []

    gray = cv2.cvtColor(page, cv2.COLOR_BGR2GRAY)
    # Denoise first — helps all subsequent pipelines
    gray = cv2.fastNlMeansDenoising(gray, h=10)

    found: dict = {}
    pipeline_used = 'none'

    for method in _QR_PIPELINES:
        processed = _preprocess(gray, method)
        results   = _decode_image(processed)
        if results:
            for r in results:
                found[r] = True
            pipeline_used = method
            break   # stop early — no need to try harder pipelines

    # Last resort: zxingcpp — handles partial/damaged QR better than pyzbar
    if not found and _HAS_ZXINGCPP:
        try:
            results = zxingcpp.read_barcodes(page)
            for r in results:
                if r.text:
                    found[r.text] = True
            if found:
                pipeline_used = 'zxingcpp'
        except Exception as e:
            write_log_error(f'zxingcpp error: {e}', 'QR', e)

    if found:
        write_log_debug(f'QR decoded via [{pipeline_used}]: {list(found.keys())}', 'QR')
    else:
        write_log_error(f'QR not found in image: {img_path}', 'QR')

    return list(found.keys())


# ─── Attendance log retrieval ─────────────────────────────────────────────────
def get_att_log_one_time(machine: ZK, machineNo: int) -> int:
    if not is_in_update_window():
        write_log_info(
            f'get_att_log_one_time: machine {machineNo} — outside update window '
            f'({update_time_from}–{update_time_to}), skip', 'ATT'
        )
        return 0
    conn = None
    count = 0
    time_begin = datetime.now()
    write_log_info(f'get_att_log_one_time: machine {machineNo}', 'ATT')
    try:
        history      = collection_history_get_att_logs.find_one({"machine": machineNo}) or {}
        last_time_db = history.get('lastTimeGetAttLogs', datetime(2020, 1, 1))
        last_time_machine = last_time_db

        conn = machine.connect()
        write_log_info(f'Connected machine {machineNo}', 'ATT')
        conn.disable_device()
        attendances = conn.get_attendance()

        for attendance in attendances:
            if attendance.timestamp > last_time_db:
                finger_id = int(attendance.user_id)
                mydict = {
                    "machineNo":   machineNo,
                    "uid":         attendance.uid,
                    "attFingerId": finger_id,
                    "empId":       emp_by_finger_id.get(finger_id, {}).get('empId', 'No Emp Id'),
                    "name":        emp_by_finger_id.get(finger_id, {}).get('name', 'No name'),
                    "timestamp":   attendance.timestamp,
                }
                collection_att_log.insert_one(mydict)
                write_log_info(
                    f'M:{machineNo} {attendance.timestamp:%d-%m-%Y %H:%M:%S} '
                    f'{mydict["attFingerId"]} {mydict["empId"]} {mydict["name"]}', 'ATT'
                )
                count += 1
                last_time_machine = attendance.timestamp

        collection_history_get_att_logs.update_one(
            {"machine": machineNo},
            {"$set": {"lastTimeGetAttLogs": last_time_machine, "lastCount": count}}
        )
    except Exception as e:
        write_log_error(f'get_att_log_one_time machine {machineNo}: {e}', 'ATT', e)
    finally:
        if conn:
            conn.enable_device()
            conn.disconnect()
        write_log_info(
            f'Machine {machineNo} => {count} records. '
            f'Total time: {datetime.now() - time_begin}', 'ATT'
        )
    return count


def live_capture_attendance(machine: ZK, machineNo: int) -> None:
    """Live capture with auto-reconnect (exponential backoff) + zombie detection."""
    reconnect_delay = RECONNECT_DELAY_MIN

    while True:
        conn         = None
        consecutive_none = 0
        connected_at = None

        try:
            write_log_info(f'Connecting machine {machineNo} for live capture...', 'LIVE')
            conn = machine.connect()
            connected_at    = datetime.now()
            reconnect_delay = RECONNECT_DELAY_MIN
            write_log_info(f'Machine {machineNo} connected. Starting live capture.', 'LIVE')

            for attendance in conn.live_capture(new_timeout=LIVE_CAPTURE_TIMEOUT):
                if attendance is None:
                    consecutive_none += 1
                    if consecutive_none >= ZOMBIE_THRESHOLD:
                        write_log_error(
                            f'Machine {machineNo}: {consecutive_none} consecutive timeouts '
                            f'({consecutive_none * LIVE_CAPTURE_TIMEOUT}s no data). Force reconnect.', 'LIVE'
                        )
                        break
                    continue

                consecutive_none = 0
                finger_id = int(attendance.user_id)
                mydict = {
                    "machineNo":   machineNo,
                    "uid":         attendance.uid,
                    "attFingerId": finger_id,
                    "empId":       emp_by_finger_id.get(finger_id, {}).get('empId', 'No Emp Id'),
                    "name":        emp_by_finger_id.get(finger_id, {}).get('name', 'No name'),
                    "timestamp":   attendance.timestamp,
                }
                collection_att_log.insert_one(mydict)
                collection_history_get_att_logs.update_one(
                    {"machine": machineNo},
                    {"$set": {"lastTimeGetAttLogs": attendance.timestamp, "lastCount": 1}}
                )
                write_log_info(
                    f'L-M:{machineNo} {attendance.timestamp:%d-%m-%Y %H:%M:%S} '
                    f'{mydict["attFingerId"]} {mydict["empId"]} {mydict["name"]}', 'LIVE'
                )

        except (KeyboardInterrupt, SystemExit):
            write_log_info(f'Machine {machineNo} live capture stopped by user.', 'LIVE')
            break

        except Exception as e:
            uptime = int((datetime.now() - connected_at).total_seconds()) if connected_at else 0
            write_log_error(
                f'Machine {machineNo} live capture error (uptime {uptime}s): {e}. '
                f'Reconnect in {reconnect_delay}s...', 'LIVE', e
            )
            if uptime > 60:
                reconnect_delay = RECONNECT_DELAY_MIN

        finally:
            if conn:
                try:
                    conn.disconnect()
                except Exception:
                    pass

        write_log_info(f'Machine {machineNo}: waiting {reconnect_delay}s before reconnect...', 'LIVE')
        time_module.sleep(reconnect_delay)
        reconnect_delay = min(reconnect_delay * RECONNECT_DELAY_MULTIPLIER, RECONNECT_DELAY_MAX)


def get_att_log(machine: ZK, machineNo: int, real_time: bool) -> None:
    get_att_log_one_time(machine, machineNo)
    if real_time:
        live_capture_attendance(machine, machineNo)
    else:
        interval = schedule_config.get('att_log_interval_minutes', ATT_LOG_INTERVAL_MINUTES)
        while True:
            time_module.sleep(interval * 60)
            get_att_log_one_time(machine, machineNo)


# ─── Device time sync ─────────────────────────────────────────────────────────
def sync_time_devices() -> None:
    write_log_info('sync_time_devices start', 'SYNC')
    for ip in ip_att_machines:
        conn = None
        try:
            machine  = ZK(ip, port=4370, timeout=5, password=0, force_udp=False, ommit_ping=False)
            conn     = machine.connect()
            now_time = datetime.now()
            write_log_info(f'Machine {ip}: current={conn.get_time()} → sync to {now_time}', 'SYNC')
            conn.set_time(now_time)
        except Exception as e:
            write_log_error(f'Sync time ERROR for {ip}: {e}', 'SYNC', e)
        finally:
            if conn:
                try:
                    conn.disconnect()
                except Exception:
                    pass


# ─── OT Register ─────────────────────────────────────────────────────────────
def ot_register_detect_qr_and_save() -> None:
    """Scan PDF files for QR codes and register OT requests into MongoDB."""
    ot_folder_path              = path_config['ot_folder']
    ot_folder_pdf_imported_path = ot_folder_path + r"\01.Imported"
    ot_folder_pdf_path          = ot_folder_path + r"\02.Pdf"
    ot_excel_file_name          = "01.OT summary.xlsx"

    ot_request_no_on_db: set = {str(doc['requestNo']) for doc in collection_ot_register.find()}

    ot_last_request_id_db = 0
    for doc in collection_ot_register.find().sort({'_id': -1}).limit(1):
        ot_last_request_id_db = doc['_id']

    pdf_files = []
    try:
        for filename in os.listdir(ot_folder_pdf_path):
            if filename.endswith('.pdf'):
                pdf_files.append(filename)
    except Exception as e:
        write_log_error(f'Error listing folder {ot_folder_pdf_path}: {e}', 'OT', e)

    for file_pdf in pdf_files:
        file_path = os.path.join(ot_folder_pdf_path, file_pdf)
        write_log_info(f'Read QR code in file: {file_path}', 'OT')
        try:
            # QR_PDF_DPI=400 for better quality on blurry scans
            temp_pages = convert_from_path(
                file_path, dpi=QR_PDF_DPI,
                output_file='qr.png', paths_only=True,
                output_folder=ot_folder_pdf_path, poppler_path=poppler_path
            )
            for temp_page in temp_pages:
                try:
                    # Use enhanced multi-pipeline decoder instead of bare pyzbar
                    qr_data_list = decode_qr_from_image(temp_page)
                    for qr in qr_data_list:
                        write_log_info(f'File: {file_pdf} QR: {qr}', 'OT')
                        parts          = qr.split(';')
                        ot_request_no  = parts[0].strip()
                        date_request   = parts[1].strip()
                        list_date_time = parts[2].strip().split(', ')
                        list_emp_id    = parts[3].strip().split(' ')
                        if ot_request_no not in ot_request_no_on_db:
                            ot_request_no_on_db.add(ot_request_no)
                            ot_last_request_id_db = qr_code_ot_register_to_db(
                                ot_last_request_id_db, ot_request_no,
                                date_request, list_date_time, list_emp_id
                            )
                            write_log_info(f'Added QR to DB: {qr}', 'OT')
                        else:
                            write_log_info(f'Duplicate QR (skip): {qr}', 'OT')
                except Exception as e:
                    write_log_error(f'detect_qr page error: {e}', 'OT', e)
        except Exception as e:
            write_log_error(f'detect_qr pdf error ({file_pdf}): {e}', 'OT', e)

        os.rename(
            os.path.join(ot_folder_pdf_path, file_pdf),
            os.path.join(ot_folder_pdf_imported_path, file_pdf)
        )
        write_log_info(f'Scanned & moved to "01.Imported": {file_pdf}', 'OT')

        for filename in os.listdir(ot_folder_pdf_path):
            if filename.endswith('.ppm'):
                os.remove(os.path.join(ot_folder_pdf_path, filename))

    try:
        ot_register_append_excel(ot_folder_path, ot_excel_file_name)
    except Exception as e:
        write_log_error(f'ot_register_append_excel error: {e}', 'OT', e)


def qr_code_ot_register_to_db(
    ot_last_request_id_db: int,
    ot_request_no: str,
    date_request: str,
    list_date_time: list,
    list_emp_id: list
) -> int:
    write_log_info(
        f'qr_code_ot_register_to_db: {ot_request_no} '
        f'dates={list_date_time} emps={list_emp_id}', 'OT'
    )
    for date_and_time in list_date_time:
        if len(date_and_time) != 20 or '19000100' in date_and_time:
            continue
        parts        = date_and_time.strip().split(' ')
        ot_date      = datetime.strptime(parts[0][:8], '%Y%m%d')
        request_date = datetime.strptime(date_request, '%Y%m%d')
        for id_no in list_emp_id:
            emp_id  = 'TIQN-' + str(id_no)
            emp_name = find_name_by_emp_id(emp_id)
            ot_last_request_id_db += 1
            mydict = {
                "_id":         ot_last_request_id_db,
                "requestNo":   ot_request_no,
                "requestDate": request_date,
                "otDate":      ot_date,
                "otTimeBegin": parts[1],
                "otTimeEnd":   parts[2],
                "empId":       emp_id,
                "name":        emp_name,
            }
            write_log_info(
                f'OT record: No={ot_request_no} Date={ot_date:%d-%m-%Y} '
                f'{parts[1]}-{parts[2]} {emp_id} {emp_name}', 'OT'
            )
            collection_ot_register.insert_one(mydict)
    return ot_last_request_id_db


def ot_register_append_excel(folder_path: str, file_name: str) -> None:
    filename              = os.path.join(folder_path, file_name)
    ot_last_request_id_db = 0
    for doc in collection_ot_register.find().sort({'_id': -1}).limit(1):
        ot_last_request_id_db = doc['_id']
    if ot_last_request_id_db == 0:
        return

    wb = load_workbook(filename=filename)
    ws = wb["data"]

    if ws.max_row == 1:
        ot_last_request_id_excel = 0
        last_row = 1
    else:
        last_row                 = ws.max_row
        ot_last_request_id_excel = int(ws.cell(last_row, 1).value)

    if ot_last_request_id_db > ot_last_request_id_excel:
        write_log_info('ot_register_append_excel: appending new records', 'OT')
        for doc in collection_ot_register.find({"_id": {"$gt": ot_last_request_id_excel}}):
            write_log_info(f'Append _id: {doc["_id"]}', 'OT')
            last_row += 1
            ws.cell(row=last_row, column=1).value = doc['_id']
            ws.cell(row=last_row, column=2).value = doc['requestNo']
            ws.cell(row=last_row, column=3).value = doc['requestDate']
            ws.cell(row=last_row, column=4).value = doc['otDate']
            ws.cell(row=last_row, column=5).value = doc['otTimeBegin']
            ws.cell(row=last_row, column=6).value = doc['otTimeEnd']
            ws.cell(row=last_row, column=7).value = doc['empId']
            ws.cell(row=last_row, column=8).value = doc['name']

        date_format = "DD-MM-YYYY"
        for row in ws.iter_rows(min_row=2):
            for col_idx in (2, 3):
                cell = row[col_idx]
                if isinstance(cell.value, datetime):
                    cell.number_format = date_format

        wb.save(filename=filename)
        wb.close()
        write_log_info(f'ot_register_append_excel done: {filename}', 'OT')


# ─── Schedule setup ───────────────────────────────────────────────────────────
def setup_schedule() -> None:
    """Register all scheduled jobs from schedule_config."""
    sc          = schedule_config
    day         = sc.get('sync_time_day', 'sunday')
    at          = sc.get('sync_time_at', '06:00')
    ot_interval = sc.get('ot_scan_interval_minutes', 10)
    getattr(schedule.every(), day).at(at).do(sync_time_devices)
    for t in sc.get('excel_sync_times', []):
        schedule.every().day.at(t).do(update_excel_to_mongoDb)
    schedule.every(ot_interval).minutes.do(ot_register_detect_qr_and_save)
    write_log_info(
        f'Schedule: sync_time={day} {at} | '
        f'excel={sc.get("excel_sync_times")} | '
        f'ot_scan={ot_interval}min | '
        f'att_log={sc.get("att_log_interval_minutes")}min', 'CONFIG'
    )
