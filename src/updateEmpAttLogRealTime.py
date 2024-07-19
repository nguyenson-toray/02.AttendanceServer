# -*- coding: utf-8 -*-
import sys
import time
from multiprocessing import Process
import os
import schedule
from zk import ZK
import pymongo
from datetime import datetime, timedelta
import pandas as pd
import cv2
from pyzbar.pyzbar import decode
from pdf2image import convert_from_path
from openpyxl import load_workbook

CWD = os.path.dirname(os.path.realpath(__file__))
ROOT_DIR = os.path.dirname(CWD)
sys.path.append(ROOT_DIR)

client = pymongo.MongoClient("mongodb://localhost:27017/")
db = client["tiqn"]
collectionEmployee = db["Employee"]
collectionAttLog = db["AttLog"]
collectionMaternityTracking = db["MaternityTracking"]
collectionHistoryGetAttLogs = db["HistoryGetAttLogs"]
collectionLastModifiedDataHR = db["LastModifiedDataHR"]
collectionOtRegister = db["OtRegister"]
emp_id_name = {}


def floor(datetimeObj):
    return datetimeObj.replace()


def excelAllInOneToMongoDb() -> bool:
    # Read data from Excel using pandas
    excelFile = r"\\fs\tiqn\03.Department\01.Operation Management\03.HR-GA\01.HR\Toray's employees information All in one.xlsx"
    lastModifiedFile = datetime.fromtimestamp(os.path.getmtime(excelFile)).replace(microsecond=0)
    lastModifiedDb = collectionLastModifiedDataHR.find_one()['aio'].replace(microsecond=0)
    needUpdate = True if lastModifiedFile > lastModifiedDb else False
    if not needUpdate:
        print(f"{excelFile} => No change from {lastModifiedDb} => pass")
        return needUpdate
    print(f"{excelFile} => Changed at {lastModifiedFile} => Need update")
    query = {'_id': 1}
    newValue = {"$set": {"aio": lastModifiedFile}}
    collectionLastModifiedDataHR.update_one(query, newValue)
    data = pd.read_excel(excelFile, keep_default_na=False, na_values='', na_filter=False)
    data.fillna("", inplace=True)
    # Convert pandas dataframe to dictionary (adjust based on your data structure)
    data_dict = data.to_dict(orient="records")
    # Loop through each data row
    count = 0
    for row in data_dict:

        # Update document in MongoDB collection based on a unique identifier (replace with your logic)
        if row["Emp Code"] == '' or row["Emp Code"] == 0 or row["Fullname"] == '' or row["Fullname"] == 0 or row[
            '_id'] == 0 or row['_id'] == '':
            # print(f'BYPASS ROW - Empty Emp Code or Fullname or _id: {row}')
            continue
        if row["Fullname"] == 'Shoji Izumi':
            # print(f'BYPASS ROW - Shoji Izumi: {row}')
            continue
        count += 1
        filter = {"_id": row["_id"]}  # Assuming "_id" is a unique identifier in your data
        # update = {"$set": row}  # Update all fields in the document
        fieldCollected = {}
        empId = 'TIQN-XXXX' if row['Emp Code'] == '' else row['Emp Code']
        name = 'No name' if row['Fullname'] == '' else row['Fullname']
        attFingerId = 0 if row['Finger Id'] == '' else row['Finger Id']
        department = '' if row['Department'] == '' else row['Department']
        section = '' if row['Section'] == '' else row['Section']
        group = '' if row['Group'] == '' else row['Group']
        lineTeam = '' if row['Line/ Team'] == '' else row['Line/ Team']
        gender = '' if row['Gender'] == '' else row['Gender']
        position = '' if row['Position'] == '' else row['Position']
        level = '' if row['Level'] == '' else row['Level']
        directIndirect = '' if row['Direct/ Indirect'] == '' else row['Direct/ Indirect']
        sewingNonSewing = '' if row['Sewing/Non sewing'] == '' else row['Sewing/Non sewing']
        supporting = '' if row['Supporting'] == '' else row['Supporting']
        dob = datetime.fromisoformat('1900-01-01') if str(type(row['DOB'])).find('datetime.datetime') == -1 else row[
            'DOB']
        joiningDate = datetime.fromisoformat('1900-01-01') if str(type(row['Joining date'])).find(
            'datetime.datetime') == -1 else row['Joining date']
        workStatus = 'Resigned' if (row['Working/Resigned'] == 0 or row['Working/Resigned'] == '') else 'Working'
        resignOn = datetime.fromisoformat('2099-01-01')
        fieldCollected.update(
            {'empId': empId, 'name': name, 'attFingerId': attFingerId, 'department': department, 'section': section,
             'group': group, 'lineTeam': lineTeam, 'gender': gender, 'position': position, 'level': level,
             'directIndirect': directIndirect, 'sewingNonSewing': sewingNonSewing, 'supporting': supporting, 'dob': dob,
             'joiningDate': joiningDate, 'workStatus': workStatus, 'resignOn': resignOn})
        update = {"$set": fieldCollected}
        # print(f"collection.update_one: {update}")
        collectionEmployee.update_one(filter, update, upsert=True)  # Upsert inserts if no document matches the filter
    print(f" excelAllInOneToMongoDb - {datetime.now()} : {count} records")
    return needUpdate


def excelMaternityToMongoDb(forceUpdate: bool):
    excelFile = r"\\fs\tiqn\03.Department\01.Operation Management\03.HR-GA\05.Nurse\5. Maternity leave\Danh sách nhân viên nữ mang thai.xlsx"
    # check if file changed
    lastModifiedFile = datetime.fromtimestamp(os.path.getmtime(excelFile)).replace(microsecond=0)
    lastModifiedDb = collectionLastModifiedDataHR.find_one()['maternity'].replace(microsecond=0)
    needUpdateMaternity = forceUpdate if forceUpdate else True if lastModifiedFile > lastModifiedDb else False
    if not needUpdateMaternity:
        print(f"{excelFile} =>No need update or no change from {lastModifiedDb} => pass")
        return
    print(f"{excelFile} =>Need to update, changed at {lastModifiedFile} => Need update")
    query = {'_id': 1}
    newValue = {"$set": {"maternity": lastModifiedFile}}
    collectionLastModifiedDataHR.update_one(query, newValue)
    # Read -thai san
    data = pd.read_excel(excelFile, sheet_name='Thai sản', keep_default_na=False, na_values='',
                         na_filter=False, skiprows=3)
    data.fillna("", inplace=True)
    # Convert pandas dataframe to dictionary (adjust based on your data structure)
    data_dict = data.to_dict(orient="records")
    # Loop through each data row
    empIdMaternity = 'TIQN-XXXX'
    for row in data_dict:
        if row['STT'] != 0 and row['STT'] != '':
            empIdMaternity = row['MSNV']
            query = {'empId': empIdMaternity}
            maternityLeaveBegin = datetime.fromisoformat('2099-01-01') if (str(type(row['NGÀY NGHỈ SINH'])).find(
                'datetime.datetime') == -1 and str(type(row['NGÀY NGHỈ SINH'])).find(
                'timestamps.Timestamp') == -1) else row['NGÀY NGHỈ SINH']
            maternityLeaveEnd = datetime.fromisoformat('2099-01-01') if (str(type(row['NGÀY QUAY LẠI'])).find(
                'datetime.datetime') == -1 and str(type(row['NGÀY QUAY LẠI'])).find(
                'timestamps.Timestamp') == -1) else row['NGÀY QUAY LẠI']
            maternityLeaveEnd = maternityLeaveEnd - timedelta(days=1)
            if (maternityLeaveBegin.year != 2099 and maternityLeaveEnd.year != 2099):
                update = {"$set": {'workStatus': 'Maternity leave', 'maternityLeaveBegin': maternityLeaveBegin,
                                   'maternityLeaveEnd': maternityLeaveEnd}}
            else:
                update = {"$set": {'workStatus': 'Maternity leave'}}
            print(f"update {empIdMaternity} to Maternity leave")
            collectionEmployee.update_one(query, update)

    ##---------mang thai--pregnant-------
    data = pd.read_excel(excelFile, sheet_name='mang thai', keep_default_na=False, na_values='',
                         na_filter=False, skiprows=3)
    data.fillna("", inplace=True)
    # Convert pandas dataframe to dictionary (adjust based on your data structure)
    data_dict = data.to_dict(orient="records")
    # Loop through each data row
    for row in data_dict:
        if row['STT'] != 0 and row['STT'] != '':
            empIdMaternity = row['MSNV']
            maternityBegin = row['NGÀY NHẬN THÔNG TIN']
            query = {'empId': empIdMaternity}
            update = {"$set": {'workStatus': 'Working pregnant', 'maternityBegin': maternityBegin}}
            collectionEmployee.update_one(query, update)
            print(f"update {empIdMaternity} to Working pregnant")
    ##---------Con nhỏ dưới 12 tháng---------
    data = pd.read_excel(excelFile, sheet_name='Con nhỏ dưới 12 tháng', keep_default_na=False, na_values='',
                         na_filter=False, skiprows=3)
    data.fillna("", inplace=True)
    # Convert pandas dataframe to dictionary (adjust based on your data structure)
    data_dict = data.to_dict(orient="records")
    # Loop through each data row
    for row in data_dict:
        if row['STT'] != 0 and row['STT'] != '':
            empIdMaternity = row['MSNV']
            maternityEnd = row['NGÀY CUỐI CÙNG THỜI GIAN NUÔI CON NHỎ']
            query = {'empId': empIdMaternity}
            update = {"$set": {'workStatus': 'Working young child', 'maternityEnd': maternityEnd}}
            print(f"update {empIdMaternity} to Working young child")
            collectionEmployee.update_one(query, update)


def excelResignToMongoDb(forceUpdate: bool):
    excelFile = r"\\fs\tiqn\03.Department\01.Operation Management\03.HR-GA\01.HR\7 Resigned list\Resigned report 1.xlsx"
    # check if file changed
    lastModifiedFile = datetime.fromtimestamp(os.path.getmtime(excelFile)).replace(microsecond=0)
    lastModifiedDb = collectionLastModifiedDataHR.find_one()['resign'].replace(microsecond=0)
    needUpdateResign = forceUpdate if forceUpdate else True if lastModifiedFile > lastModifiedDb else False
    if not needUpdateResign:
        print(f"{excelFile} =>No need update or No change from {lastModifiedDb} => pass")
        return forceUpdate
    print(f"{excelFile} =>Need to update, Changed at {lastModifiedFile} => Need update")
    query = {'_id': 1}
    newValue = {"$set": {"resign": lastModifiedFile}}
    collectionLastModifiedDataHR.update_one(query, newValue)
    data = pd.read_excel(excelFile, sheet_name='QD', keep_default_na=False, na_values='',
                         na_filter=False)
    data.fillna("", inplace=True)
    # Convert pandas dataframe to dictionary (adjust based on your data structure)
    data_dict = data.to_dict(orient="records")
    # Loop through each data row
    for row in data_dict:
        if row['Số QĐ'] != 0 and row['Số QĐ'] != '':
            empIdResigned = row['MSNV']
            resignDate = datetime.fromisoformat('2099-01-01') if (str(type(row['Ngày nghỉ việc'])).find(
                'datetime.datetime') == -1 and str(type(row['Ngày nghỉ việc'])).find(
                'timestamps.Timestamp') == -1) else row['Ngày nghỉ việc']

            if resignDate.year > 1900 and resignDate < datetime.now():
                query = {'empId': empIdResigned}
                update = {"$set": {'workStatus': 'Resigned', 'resignOn': resignDate}}
                print(f"update {empIdResigned} to resigned on {resignDate}")
                collectionEmployee.update_one(query, update)
    return forceUpdate


def get_att_log_one_time(machine: ZK, machineNo: int) -> int:
    conn = None
    count = 0
    timeBeginGetLogs = datetime.now()
    try:

        history = {}
        for his in collectionHistoryGetAttLogs.find():
            if (int(his['machine']) == machineNo):
                history = his
        lastTime = history['lastTimeGetAttLogs']
        conn = machine.connect()
        print(f"Connecting machine : {machine.get_network_params()['ip']}")
        # disable device, this method ensures no activity on the device while the process is run
        conn.disable_device()
        # another commands will be here!
        # Get attendances (will return list of Attendance object)
        attendances = conn.get_attendance()
        for attendance in attendances:
            if (attendance.timestamp > lastTime):
                mydict = {"machineNo": machineNo, "uid": attendance.uid, "attFingerId": int(attendance.user_id),
                          "empId": 'No Emp Id', "name": 'No name',
                          "timestamp": attendance.timestamp}
                for emp in collectionEmployee.find():
                    if (int(emp['attFingerId']) == int(attendance.user_id)):
                        mydict = {"machineNo": machineNo, "uid": attendance.uid, "attFingerId": int(attendance.user_id),
                                  "empId": emp['empId'], "name": emp['name'],
                                  "timestamp": attendance.timestamp}
                        break
                collectionAttLog.insert_one(mydict)
                count += 1
        myquery = {"machine": machineNo}
        newvalue = {"$set": {"lastTimeGetAttLogs": datetime.now(), "lastCount": count}}
        collectionHistoryGetAttLogs.update_one(myquery, newvalue)
        conn.enable_device()
    except:
        print('except')
    finally:
        if conn:
            conn.disconnect()
        print(f"Machine {machineNo} => {count} records. Total time: {datetime.now() - timeBeginGetLogs}")


def live_capture_attendance(machine: ZK, machineNo) -> None:
    conn = None
    try:
        conn = machine.connect()
        print(f"Live capture attendance: {machine.get_network_params()['ip']}")
        for attendance in conn.live_capture():
            if attendance is None:
                pass
            else:
                mydict = {"machineNo": machineNo, "uid": attendance.uid, "attFingerId": int(attendance.user_id),
                          "empId": 'No Emp Id', "name": 'No name',
                          "timestamp": attendance.timestamp}
                for emp in collectionEmployee.find():
                    if (int(emp['attFingerId']) == int(attendance.user_id)):
                        mydict = {"machineNo": machineNo, "uid": attendance.uid,
                                  "attFingerId": int(attendance.user_id),
                                  "empId": emp['empId'], "name": emp['name'],
                                  "timestamp": attendance.timestamp}
                        break
                print(mydict)
                collectionAttLog.insert_one(mydict)
                myquery = {"machine": machineNo}
                newvalue = {"$set": {"lastTimeGetAttLogs": datetime.now(), "lastCount": 1}}
                collectionHistoryGetAttLogs.update_one(myquery, newvalue)

    except Exception as e:
        print("Process terminate : {}".format(e))
    finally:
        if conn:
            conn.disconnect()


def update_excel_to_mongoDb():
    needUpdate = excelAllInOneToMongoDb()
    excelMaternityToMongoDb(needUpdate)
    excelResignToMongoDb(needUpdate)


def get_emp_id_name():
    emp_id_name.clear()
    for emp in collectionEmployee.find():
        emp_id_name[emp['empId']] = emp['name']


def ot_register_append_excel(folder_path: str, file_name):
    filename = os.path.join(folder_path, file_name)
    # if open(filename, 'r'):
    #     print(f'ot_register_append_excel : file can not OPEN or ERROR: {filename}')
    #     return
    # ot_last_request_id_db = collectionLastModifiedDataHR.find_one()['otRequestId']
    ot_last_request_id_db = 0
    ot_last_ot_register_document = collectionOtRegister.find().sort({'_id': -1}).limit(1)
    for doc in ot_last_ot_register_document:
        ot_last_request_id_db = doc['_id']
    if ot_last_request_id_db == 0:
        return
    print(F'Append OT data from db to excel')
    # Open the existing workbook
    wb = load_workbook(filename=filename)
    sheet_name = "data"  # Replace with your sheet name
    # Select the sheet you want to append data to (adjust sheet name)
    ws = wb[sheet_name]

    if ws.max_row == 1:
        ot_last_request_id_excel = 0
        last_row = 1
    else:
        last_row = ws.max_row
        ot_last_request_id_excel = int(ws.cell(last_row, 1).value)
    if ot_last_request_id_db > ot_last_request_id_excel:
        query = {"_id": {"$gt": ot_last_request_id_excel}}
        docs = collectionOtRegister.find(query)
        for doc in docs:
            print(f"  =>  Append row {last_row} : value: {doc}")
            last_row += 1
            ws.cell(row=last_row, column=1).value = doc['_id']
            ws.cell(row=last_row, column=2).value = doc['requestNo']
            ws.cell(row=last_row, column=3).value = doc['requestDate']
            ws.cell(row=last_row, column=4).value = doc['otDate']
            ws.cell(row=last_row, column=5).value = doc['otTimeBegin']
            ws.cell(row=last_row, column=6).value = doc['otTimeEnd']
            ws.cell(row=last_row, column=7).value = doc['empId']
            ws.cell(row=last_row, column=8).value = doc['name']
        # Apply date formatting to the entire column
        date_format = "DD-MM-YYYY"
        for row in ws.iter_rows(min_row=2):  # Skip the first row (headers)
            cell = row[2]  # Access cell by column index (0-based)
            if isinstance(cell.value, datetime):  # Check if value is a date object
                cell.number_format = date_format
            else:
                pass
            cell = row[3]  # Access cell by column index (0-based)
            if isinstance(cell.value, datetime):  # Check if value is a date object
                cell.number_format = date_format
            else:
                pass
        # Save the modified workbook
        wb.save(filename=filename)
        wb.close()
        print(f"  => ot_register_append_excel : Data appended to: {filename}")


def qr_code_ot_register_to_db(ot_last_request_id_db: int, ot_request_no: str, date_request: str, list_date_time: list,
                              list_emp_id: list) -> int:
    print(f'qr_code_ot_register_to_db: ot_request_no :{ot_request_no}')
    for date_and_time in list_date_time:
        if not len(date_and_time) == 20 or date_and_time.find('19000100') >= 0:
            print(f'{date_and_time}: is wrong => pass')
            continue
        date_ot_str = date_and_time.strip().split(' ')[0]
        ot_time_begin_str = date_and_time.strip().split(' ')[1]
        ot_time_end_str = date_and_time.strip().split(' ')[2]
        ot_date = datetime.strptime(date_ot_str[0:8], '%Y%m%d')
        request_date = datetime.strptime(date_request, '%Y%m%d')
        for id_no in list_emp_id:
            emp_id = 'TIQN-' + str(id_no)
            emp_name = emp_id_name[emp_id]
            ot_last_request_id_db += 1
            mydict = {"_id": ot_last_request_id_db, "requestNo": ot_request_no, "requestDate": request_date,
                      "otDate": ot_date, 'otTimeBegin': ot_time_begin_str, 'otTimeEnd': ot_time_end_str,
                      "empId": emp_id, "name": emp_name}
            collectionOtRegister.insert_one(mydict)
    return ot_last_request_id_db


def ot_register_detect_qr_and_save():
    begin = datetime.now()
    ot_folder_path = r"\\fs\tiqn\03.Department\01.Operation Management\03.HR-GA\01.HR\20.OT request"
    ot_folder_pdf_path = ot_folder_path + r"\02.Pdf"
    ot_excel_file_name = "01.OT summary.xlsx"
    my_poppler_path = r'C:\Program Files\poppler-24.02.0\Library\bin'
    # ot_last_request_id_db = collectionLastModifiedDataHR.find_one()['otRequestId']
    ot_last_request_id_db = 0
    ot_last_ot_register_document = collectionOtRegister.find().sort({'_id': -1}).limit(1)
    for doc in ot_last_ot_register_document:
        ot_last_request_id_db = doc['_id']
    # ot_last_form_request_time=None
    for doc in collectionLastModifiedDataHR.find().limit(1):
        ot_last_form_request_time = doc['otFormRequestTime']
    # print(f'detect_qr : database : last otRequestId = {ot_last_request_id_db}      last otFormRequestTime: {ot_last_form_request_time}')
    list_qr = []
    pdf_paths = []
    try:
        # Use glob with wildcard for efficient PDF search
        for filename in os.listdir(ot_folder_pdf_path):
            if filename.endswith('.pdf'):  # and filename.startswith("202"):
                pdf_paths.append(os.path.join(ot_folder_pdf_path, filename))
        # Sort PDFs by last modified time (descending) using reverse=True
        pdf_paths.sort(key=os.path.getmtime)
        # print(f"pdf_paths:{pdf_paths}")
    except Exception as e:
        print(f"Error processing folder {ot_folder_pdf_path}: {e}")
    for file_path in pdf_paths:
        file_last_modified = datetime.fromtimestamp(os.path.getmtime(file_path))
        print(f'File: {file_path} => file_last_modified : {file_last_modified}')
        list_qr = []
        diff_time = file_last_modified - ot_last_form_request_time
        if diff_time.total_seconds() > 1:
            try:
                temp_pages = convert_from_path(file_path, dpi=300, output_file='qr.png', paths_only=True,
                                               output_folder=ot_folder_pdf_path, poppler_path=my_poppler_path)
                for temp_page in temp_pages:
                    try:
                        page = cv2.imread(temp_page)
                        gray = cv2.cvtColor(page, cv2.COLOR_BGR2GRAY)  # Convert to grayscale for better detection
                        # Detect QR codes
                        qr_codes = decode(gray)
                        # Extract data from QR codes
                        for qr_code in qr_codes:
                            qr = str(qr_code.data.decode('utf-8'))
                            print(f'QR : {qr}')
                            ot_request_no = qr.split(';')[0].strip()
                            date_request = qr.split(';')[1].strip()
                            list_date_time = qr.split(';')[2].strip().split(', ')
                            list_emp_id = qr.split(';')[3].strip().split(' ')
                            ot_last_request_time = file_last_modified
                            ot_last_request_id_db = qr_code_ot_register_to_db(ot_last_request_id_db, ot_request_no,
                                                                              date_request,
                                                                              list_date_time, list_emp_id)
                            query = {'_id': 1}
                            update = {
                                "$set": {'otFormRequestTime': ot_last_request_time,
                                         # "otRequestId": ot_last_request_id_db
                                         }}
                            collectionLastModifiedDataHR.update_one(query, update)

                    except Exception as e:
                        print(f"detect_qr : Error 1 :{e}")

            except Exception as e:
                print(f"detect_qr : Error 2 :{e}")
        else:
            print(f'  => Pass')
    try:
        ot_register_append_excel(ot_folder_path, ot_excel_file_name)
    except Exception as e:
        print(f"detect_qr : Error 3 :{e}")

    for filename in os.listdir(ot_folder_pdf_path):
        file_path = os.path.join(ot_folder_pdf_path, filename)
        if filename.endswith('.ppm'):
            os.remove(file_path)
    end = datetime.now()
    time = end - begin
    print(f' Total time total_seconds : {time.total_seconds()}')


def sync_time_devices(machines: [ZK]):
    try:
        for machine in machines:
            conn = machine.connect()
            print(f"Machine :  {conn.get_network_params()}")
            print(f" Current time : {conn.get_time()}")
            time = datetime.now()
            print(f' Sync time to : {time}')
            conn.set_time(time)
    except Exception as e:
        print(" Sync time ERROR : {}".format(e))


if __name__ == "__main__":  # confirms that the code is under main function

    machine1 = ZK('192.168.1.31', port=4370, timeout=5, password=0, force_udp=False, ommit_ping=False)
    machine2 = ZK('192.168.1.32', port=4370, timeout=5, password=0, force_udp=False, ommit_ping=False)
    machine3 = ZK('192.168.1.33', port=4370, timeout=5, password=0, force_udp=False, ommit_ping=False)
    machine4 = ZK('192.168.1.34', port=4370, timeout=5, password=0, force_udp=False, ommit_ping=False)
    attMachines = [machine1, machine2, machine3, machine4]

    run_one_time = False
    if run_one_time:
        sync_time_devices(attMachines)
        # get_emp_id_name()
        # update_excel_to_mongoDb()
        # ot_register_detect_qr_and_save()
        # machineNo = 0
        # for machine in attMachines:
        #     machineNo += 1
        #     Process(target=get_att_log_one_time, args=(machine, machineNo)).start()
    else:
        get_emp_id_name()
        update_excel_to_mongoDb()
        ot_register_detect_qr_and_save()
        machineNo = 0
        for machine in attMachines:
            machineNo += 1
            Process(target=live_capture_attendance, args=(machine, machineNo)).start()
            # Create a new process
        schedule.every().minute.do(update_excel_to_mongoDb)
        schedule.every().minute.do(ot_register_detect_qr_and_save)
        schedule.every().day.at("00:00").do(sync_time_devices)
        while True:
            schedule.run_pending()
            time.sleep(1)
